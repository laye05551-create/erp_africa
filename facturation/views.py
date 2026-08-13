from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from entreprises.models import MembreEntreprise, Entreprise
from .models import Client, Facture, LigneFacture
from stocks.models import Produit
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def get_entreprise(request):
    if request.user.is_superuser:
        return Entreprise.objects.order_by('id').first()
    try:
        membre = MembreEntreprise.objects.get(user=request.user)
        if not membre.actif:
            return None
        return membre.entreprise
    except MembreEntreprise.DoesNotExist:
        return None


def get_role(request):
    if request.user.is_superuser:
        return 'AD'
    try:
        membre = MembreEntreprise.objects.get(user=request.user)
        return membre.role
    except MembreEntreprise.DoesNotExist:
        return None


@login_required
def liste_clients(request):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    recherche = request.GET.get('q', '')
    clients = Client.objects.filter(entreprise=entreprise)
    if recherche:
        clients = clients.filter(nom__icontains=recherche) | clients.filter(telephone__icontains=recherche) | clients.filter(email__icontains=recherche)
    return render(request, 'facturation/clients.html', {
        'clients': clients,
        'entreprise': entreprise,
        'recherche': recherche
    })


@login_required
def ajouter_client(request):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    if request.method == 'POST':
        Client.objects.create(
            entreprise=entreprise,
            nom=request.POST.get('nom'),
            telephone=request.POST.get('telephone'),
            email=request.POST.get('email'),
            adresse=request.POST.get('adresse'),
            ninea=request.POST.get('ninea'),
        )
        messages.success(request, 'Client ajouté avec succès !')
        return redirect('/clients/')
    return render(request, 'facturation/ajouter_client.html', {'entreprise': entreprise})


@login_required
def modifier_client(request, client_id):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    client = get_object_or_404(Client, id=client_id, entreprise=entreprise)
    if request.method == 'POST':
        client.nom = request.POST.get('nom')
        client.telephone = request.POST.get('telephone')
        client.email = request.POST.get('email')
        client.adresse = request.POST.get('adresse')
        client.ninea = request.POST.get('ninea')
        client.save()
        messages.success(request, 'Client modifié avec succès !')
        return redirect('/clients/')
    return render(request, 'facturation/modifier_client.html', {'client': client, 'entreprise': entreprise})


@login_required
def supprimer_client(request, client_id):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    client = get_object_or_404(Client, id=client_id, entreprise=entreprise)
    if request.method == 'POST':
        client.delete()
        messages.success(request, 'Client supprimé !')
        return redirect('/clients/')
    return render(request, 'facturation/supprimer_client.html', {'client': client, 'entreprise': entreprise})


@login_required
def liste_factures(request):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    role = get_role(request)
    if role not in ['AD', 'CM', 'CO']:
        messages.error(request, 'Accès refusé.')
        return redirect('/dashboard/')
    recherche = request.GET.get('q', '')
    factures = Facture.objects.filter(entreprise=entreprise, est_supprime=False)
    if recherche:
        factures = factures.filter(numero__icontains=recherche) | factures.filter(client__nom__icontains=recherche)
    return render(request, 'facturation/factures.html', {
        'factures': factures,
        'entreprise': entreprise,
        'role': role,
        'recherche': recherche
    })


@login_required
def ajouter_facture(request):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    clients = Client.objects.filter(entreprise=entreprise)
    produits = Produit.objects.filter(entreprise=entreprise)
    if request.method == 'POST':
        client_id = request.POST.get('client')
        numero = request.POST.get('numero')
        date_echeance = request.POST.get('date_echeance')
        client = get_object_or_404(Client, id=client_id)
        facture = Facture.objects.create(
            entreprise=entreprise,
            client=client,
            numero=numero,
            date_echeance=date_echeance,
            statut='BR',
        )
        produit_ids = request.POST.getlist('produit')
        quantites = request.POST.getlist('quantite')
        prix = request.POST.getlist('prix_unitaire')
        for i in range(len(produit_ids)):
            if produit_ids[i]:
                produit = get_object_or_404(Produit, id=produit_ids[i])
                LigneFacture.objects.create(
                    facture=facture,
                    produit=produit,
                    quantite=quantites[i],
                    prix_unitaire=prix[i],
                )
        messages.success(request, 'Facture créée avec succès !')
        return redirect('/factures/')
    return render(request, 'facturation/ajouter_facture.html', {
        'clients': clients,
        'produits': produits,
        'entreprise': entreprise
    })


@login_required
def modifier_statut_facture(request, facture_id):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    facture = get_object_or_404(Facture, id=facture_id, entreprise=entreprise)
    if request.method == 'POST':
        facture.statut = request.POST.get('statut')
        facture.save()
        messages.success(request, 'Statut mis à jour !')
        return redirect('/factures/')
    return render(request, 'facturation/modifier_statut.html', {'facture': facture, 'entreprise': entreprise})


@login_required
def supprimer_facture(request, facture_id):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    facture = get_object_or_404(Facture, id=facture_id, entreprise=entreprise)
    if request.method == 'POST':
        from django.utils import timezone
        facture.est_supprime = True
        facture.date_suppression = timezone.now()
        facture.save()
        messages.success(request, 'Facture archivée.')
        return redirect('/factures/')
    return render(request, 'facturation/supprimer_facture.html', {'facture': facture, 'entreprise': entreprise})


@login_required
def generer_pdf(request, facture_id):
    entreprise = get_entreprise(request)
    facture = get_object_or_404(Facture, id=facture_id, entreprise=entreprise)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Facture_{facture.numero}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle('titre', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#1a1a2e'), spaceAfter=20)
    elements.append(Paragraph(f"🌍 {entreprise.nom}", titre_style))
    elements.append(Paragraph(f"<b>FACTURE N° {facture.numero}</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.5*cm))
    infos = [
        ['Date émission:', str(facture.date_emission)],
        ['Date échéance:', str(facture.date_echeance)],
        ['Statut:', facture.get_statut_display()],
    ]
    table_infos = Table(infos, colWidths=[4*cm, 8*cm])
    table_infos.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('TEXTCOLOR', (0,0), (0,-1), colors.grey),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(table_infos)
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("<b>CLIENT</b>", styles['Heading3']))
    elements.append(Paragraph(facture.client.nom, styles['Normal']))
    elements.append(Paragraph(facture.client.telephone, styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    data = [['Produit', 'Quantité', 'Prix Unitaire', 'Total']]
    for ligne in facture.lignes.all():
        data.append([ligne.produit.nom, str(ligne.quantite), f"{ligne.prix_unitaire:,.0f} FCFA", f"{ligne.total:,.0f} FCFA"])
    data.append(['', '', 'Total HT:', f"{facture.total_ht:,.0f} FCFA"])
    data.append(['', '', 'TVA (18%):', f"{facture.tva:,.0f} FCFA"])
    data.append(['', '', 'TOTAL TTC:', f"{facture.total_ttc:,.0f} FCFA"])
    table = Table(data, colWidths=[8*cm, 3*cm, 4*cm, 4*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-4), 0.5, colors.grey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#E1F5EE')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(f"Merci pour votre confiance — {entreprise.nom} | {entreprise.ville}, Sénégal", styles['Normal']))
    doc.build(elements)
    return response


@login_required
def exporter_factures_excel(request):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures"
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    headers = ['Numéro', 'Client', 'Date Emission', 'Date Echéance', 'Total HT', 'TVA 18%', 'Total TTC', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 12
    factures = Facture.objects.filter(entreprise=entreprise, est_supprime=False)
    statuts = {'BR': 'Brouillon', 'EN': 'Envoyée', 'PY': 'Payée', 'AN': 'Annulée'}
    for row, facture in enumerate(factures, 2):
        ws.cell(row=row, column=1, value=facture.numero)
        ws.cell(row=row, column=2, value=facture.client.nom)
        ws.cell(row=row, column=3, value=str(facture.date_emission))
        ws.cell(row=row, column=4, value=str(facture.date_echeance))
        ws.cell(row=row, column=5, value=float(facture.total_ht))
        ws.cell(row=row, column=6, value=float(facture.tva))
        ws.cell(row=row, column=7, value=float(facture.total_ttc))
        ws.cell(row=row, column=8, value=statuts.get(facture.statut, facture.statut))
        if facture.statut == 'PY':
            fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        elif facture.statut == 'AN':
            fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = fill
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Factures_{entreprise.nom}.xlsx"'
    wb.save(response)
    return response


@login_required
def exporter_clients_excel(request):
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    headers = ['Nom', 'Téléphone', 'Email', 'Adresse', 'NINEA']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    clients = Client.objects.filter(entreprise=entreprise)
    for row, client in enumerate(clients, 2):
        ws.cell(row=row, column=1, value=client.nom)
        ws.cell(row=row, column=2, value=client.telephone)
        ws.cell(row=row, column=3, value=client.email)
        ws.cell(row=row, column=4, value=client.adresse)
        ws.cell(row=row, column=5, value=client.ninea)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Clients_{entreprise.nom}.xlsx"'
    wb.save(response)
    return response

@login_required
def detail_facture(request, facture_id):
    """Afficher les détails complets d'une facture"""
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    
    facture = get_object_or_404(Facture, id=facture_id, entreprise=entreprise)
    
    # Vérifier les droits : seul AD, CM, CO peuvent voir
    role = get_role(request)
    if role not in ['AD', 'CM', 'CO']:
        messages.error(request, 'Accès refusé.')
        return redirect('/dashboard/')
    
    return render(request, 'facturation/detail_facture.html', {
        'facture': facture,
        'entreprise': entreprise,
    })


@login_required
def modifier_facture(request, facture_id):
    """Modifier une facture existante (client, lignes, dates)"""
    entreprise = get_entreprise(request)
    if not entreprise:
        return redirect('/')
    
    facture = get_object_or_404(Facture, id=facture_id, entreprise=entreprise)
    
    # Vérifier les droits : seul AD et CM peuvent modifier
    role = get_role(request)
    if role not in ['AD', 'CM']:
        messages.error(request, 'Seuls les administrateurs et comptables peuvent modifier.')
        return redirect('/factures/')
    
    # On ne peut modifier que si la facture est en brouillon
    if facture.statut != 'BR':
        messages.error(request, 'On ne peut modifier que les factures en brouillon.')
        return redirect(f'/factures/detail/{facture_id}/')
    
    clients = Client.objects.filter(entreprise=entreprise)
    produits = Produit.objects.filter(entreprise=entreprise)
    
    if request.method == 'POST':
        # Mise à jour du client et des dates
        client_id = request.POST.get('client')
        numero = request.POST.get('numero')
        date_echeance = request.POST.get('date_echeance')
        
        try:
            client = Client.objects.get(id=client_id, entreprise=entreprise)
            facture.client = client
            facture.numero = numero
            facture.date_echeance = date_echeance
            facture.save()
        except Client.DoesNotExist:
            messages.error(request, 'Client invalide.')
            return redirect(f'/factures/modifier/{facture_id}/')
        
        # Suppression des anciennes lignes
        facture.lignes.all().delete()
        
        # Ajout des nouvelles lignes
        produit_ids = request.POST.getlist('produit')
        quantites = request.POST.getlist('quantite')
        prix = request.POST.getlist('prix_unitaire')
        
        for i in range(len(produit_ids)):
            if produit_ids[i]:
                try:
                    produit = Produit.objects.get(id=produit_ids[i], entreprise=entreprise)
                    LigneFacture.objects.create(
                        facture=facture,
                        produit=produit,
                        quantite=int(quantites[i]) if quantites[i] else 1,
                        prix_unitaire=float(prix[i]) if prix[i] else 0,
                    )
                except (Produit.DoesNotExist, ValueError):
                    continue
        
        messages.success(request, 'Facture modifiée avec succès !')
        return redirect(f'/factures/detail/{facture_id}/')
    
    return render(request, 'facturation/modifier_facture.html', {
        'facture': facture,
        'clients': clients,
        'produits': produits,
        'entreprise': entreprise
    })